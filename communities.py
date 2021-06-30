"""Downloads and converts community data into JSON, and deals with requests."""

import bottle
from datetime import datetime
from gd_download import download_file_from_gd
import json
from openpyxl import load_workbook

# settings
gd_id = "1jzH64LvFQ-UsDibXO0MOtvjbL2CvnV3N"
excel_file = "communities.xlsx"
output_file = "communities.json"
cols = ["id", "name", "rate", "count", "ltchrh_rate", "ltchrh_cases", "non_ltchrh_rate", "non_ltchrh_cases", "recent_rate", "recent_cases", "recent_ltchrh_rate", "recent_ltchrh_cases", "recent_non_ltchrh_rate", "recent_non_ltchrh_cases"]

def update():
    """Updates community data."""

    download_file_from_gd(gd_id, excel_file)

    wb = load_workbook("communities.xlsx")

    all_cases = wb["All Cases and Rates by Neighbou"]
    all_ltchrh = wb["LTCH-RH Cases and Rates by Neig"]
    recent_cases = wb["Recent Cases and Rates by Neigh"]
    recent_ltchrh = wb["Recent LTCH RH Cases and Rates "]

    # initial data (with all cases)
    communities = [dict(zip(cols, row + ((0,) * (len(cols) - 4)))) for row in all_cases.iter_rows(min_row=2, values_only=True)]

    # 2nd workbook (all ltch/rh cases)
    for row in all_ltchrh.iter_rows(min_row=2, values_only=True):
        index = next(i for i, j in enumerate(communities) if j["id"] == row[0])
        if row[2] == "LTCH RH Residents":
            communities[index][cols[4]] = row[3] # ltch/rh rate
            communities[index][cols[5]] = row[4] # ltch/rh cases
        else:
            communities[index][cols[6]] = row[3] # non ltch/rh rate
            communities[index][cols[7]] = row[4] # non ltch/rh cases
    
    # 3rd workbook (recent cases)
    for row in recent_cases.iter_rows(min_row=2, values_only=True):
        index = next(i for i, j in enumerate(communities) if j["id"] == row[0])
        communities[index][cols[8]] = row[2] # recent rate
        communities[index][cols[9]] = row[3] # recent cases
    
    # 4th workbook (recent ltch/rh cases)
    for row in recent_ltchrh.iter_rows(min_row=2, values_only=True):
        index = next(i for i, j in enumerate(communities) if j["id"] == row[0])
        if row[2] == "LTCH RH Residents":
            communities[index][cols[10]] = row[3] # recent ltch/rh rate
            communities[index][cols[11]] = row[4] # recent ltch/rh cases
        else:
            communities[index][cols[12]] = row[3] # recent non ltch/rh rate
            communities[index][cols[13]] = row[4] # recent non ltch/rh cases

    # as of
    data_date = wb["Data Note"]["A2"].value[11:]
    formatted_date = datetime.strptime(data_date, "%B %d, %Y").strftime("%m/%d/%Y")

    full_data = {"data": communities, "as_of": formatted_date}
    
    with open(output_file, "w") as f:
        json.dump(full_data, f)


@bottle.get("/communities")
@bottle.post("/communities")
def api():
    """Returns communities data that matches parameters."""

    with open(output_file, "r") as f:
        community_data = json.load(f)
        for param, value in bottle.request.params.items():
            if param in cols:
                community_data["data"] = list(filter(lambda x: str(x[param]) == value, community_data["data"]))
    return {"result": community_data, "success": True}