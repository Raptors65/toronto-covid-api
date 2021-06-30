"""Converts data from the daily status Excel file into JSON and deals with requests."""

import bottle
from datetime import datetime
from daily_cases import excel_file
import json
from openpyxl import load_workbook

# settings
cols = ["indicator", "cumulative", "newly_reported"]
output_file = "daily_status.json"

def update():
    """Updates daily report of case status."""

    # loads the downloaded excel file (from daily_cases.py)
    wb = load_workbook(excel_file)
    daily_status_data = wb["Status"]
    status = []
    for row in daily_status_data.iter_rows(min_row=2, values_only=True):
        status.append(dict(zip(cols, row)))
    
    # "as of" date
    data_date = wb["Data Note"]["A2"].value[11:]
    formatted_date = datetime.strptime(data_date, "%B %d, %Y").strftime("%m/%d/%Y")

    # combining the data and the date
    full_data = {"data": status, "as_of": formatted_date}

    # storing in JSON file
    with open(output_file, "w") as f:
        json.dump(full_data, f)

@bottle.get("/daily_status")
@bottle.post("/daily_status")
def api():
    """Returns daily COVID-19 status indicators that match parameters."""

    with open(output_file, "r") as f:
        daily_status = json.load(f)
        # filtering out data
        for param, value in bottle.request.params.items():
            if param in cols:
                daily_status["data"] = list(filter(lambda x: str(x[param]) == value, daily_status["data"]))
    return {"result": daily_status, "success": True}