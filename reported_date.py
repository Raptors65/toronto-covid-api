"""Converts reported date data from the daily status Excel file into JSON and deals with requests."""

import bottle
from datetime import datetime
from daily_cases import excel_file
import json
from openpyxl import load_workbook

# settings
cols_normal = ["date", "recovered_cases", "active_cases", "deceased_cases"]
cols_voc = ["confirmed_voc", "screened_positive_voc"]
output_file = "reported_date.json"

def update():
    """Updates data on cases by reported date."""

    # loads the downloaded excel file (from daily_cases.py)
    wb = load_workbook(excel_file)
    reported_date_data = wb["Cases by Reported Date"]
    reported_data = []
    for row in reported_date_data.iter_rows(min_row=2, values_only=True):
        # date is a datetime object so it must be converted to str
        reported_data.append(dict([(cols_normal[0], row[0].strftime("%m/%d/%Y")), *zip(cols_normal[1:], row[1:])]))
    
    # combining the VOC data
    reported_voc_data = wb["Cases by VOC and Reported Date"]
    # iterator to loop through the rows of the VOC data
    voc_rows = reported_voc_data.iter_rows(min_row=2, values_only=True)
    # getting the first row of the VOC data
    voc_row = next(voc_rows)
    voc_date = voc_row[0].strftime("%m/%d/%Y")
    # looping through the reported date data
    for i, date_data in enumerate(reported_data):
        # checking if the date is the same as the VOC date
        if voc_date == date_data[cols_normal[0]]:
            # combining the data, excluding the date since that is already in the data
            reported_data[i].update(dict(zip(cols_voc, voc_row[1:])))
            # moving on to the next VOC date
            try:
                voc_row = next(voc_rows)
                voc_date = voc_row[0].strftime("%m/%d/%Y")
            except StopIteration: # if there is no more VOC data
                # then fill the rest of the dates with 0s for VOCs
                for date_data in reported_data[i+1:]:
                    date_data.update(dict(zip(cols_voc, [0] * len(cols_voc))))

        # if there is no VOC data for this date (because there could be case but not VOC data for a specific date)
        else:
            # then we can assume there were 0 VOCs
            reported_data[i].update(dict(zip(cols_voc, [0] * len(cols_voc))))
    
    # "as of" date
    data_date = wb["Data Note"]["A2"].value[11:]
    formatted_date = datetime.strptime(data_date, "%B %d, %Y").strftime("%m/%d/%Y")

    # combining the data and the date
    full_data = {"data": reported_data, "as_of": formatted_date}

    # storing in JSON file
    with open(output_file, "w") as f:
        json.dump(full_data, f)

@bottle.get("/reported_date")
@bottle.post("/reported_date")
def api():
    """Returns COVID-19 data by reported date that matches parameters."""

    with open(output_file, "r") as f:
        reported_date = json.load(f)
        # filtering out data
        for param, value in bottle.request.params.items():
            if param in cols_normal:
                reported_date["data"] = list(filter(lambda x: str(x[param]) == value, reported_date["data"]))
    return {"result": reported_date, "success": True}